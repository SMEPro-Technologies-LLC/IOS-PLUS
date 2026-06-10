#!/usr/bin/env python3
"""
convert_matrix_seed.py — SMEPro COS+ Universal Compliance Decoding Matrix
xlsx -> database seed artifacts.

Outputs (to --out-dir):
  uco_nodes_seed.json        canonical node records (the source of truth)
  R__seed_uco_nodes.sql      Flyway REPEATABLE migration, idempotent upserts
  firecrawl_monitors.json    (--emit-monitors) monitor_create payloads for
                             high-risk federal nodes, keyed by UCO_NODE_ID
  seed_validation_report.md  what was parsed, fixed, warned, rejected

Design notes:
  * Repeatable migration (R__) rather than versioned: re-running after a
    matrix update upserts changed rows instead of requiring a new V#.
    Flyway re-applies it whenever the file checksum changes.
  * SCHEMA_MAP centralizes target column names — align these with the
    actual DDL in V4__uco_amendment.sql / V6__seed_crosswalk.sql before
    first run; nothing else in the script needs to change.
  * Known data defects are corrected via NODE_PATCHES (explicit, auditable),
    never silently.

Usage:
  python convert_matrix_seed.py MATRIX.xlsx --out-dir seed/ --emit-monitors
"""

import argparse
import json
import re
import sys
from collections import Counter
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

# --------------------------------------------------------------------------
# Target schema mapping — EDIT to match V4/V6 DDL column names.
# key = canonical field used in this script; value = DB column name.
SCHEMA_MAP = {
    "table": "uco_nodes",
    "conflict_key": "uco_node_id",
    "columns": {
        "uco_node_id": "uco_node_id",
        "sector_partition": "sector_partition",
        "broad_industry": "broad_industry",
        "industry_subtype": "industry_subtype",
        "activity": "activity_process",
        "jurisdiction_level": "jurisdiction_level",
        "governing_agency": "governing_agency",
        "regulation_name": "regulation_name",
        "citation": "citation",
        "report_form_name": "report_form_name",
        "form_code": "form_code",
        "filing_frequency": "filing_frequency",
        "due_dates": "due_dates",
        "business_segment": "business_segment",
        "penalties": "penalties",
        "cip": "code_cip",
        "sic": "code_sic",
        "naics": "code_naics",
        "soc": "code_soc",
        "isic": "code_isic",
        "hs_hts": "code_hs_hts",
        "notes": "notes",
        "ontology_level": "ontology_level",
        "compliance_chain_ref": "compliance_chain_ref",
        "operating_segment": "operating_segment",
        "responsible_role": "responsible_role",
        "enforcement_type": "enforcement_type",
        "risk_weight": "risk_weight",
        "ybr_gate": "ybr_gate",
        "policy_action": "policy_action",
        "last_updated": "last_updated",
    },
}

# Source header -> canonical field
HEADER_MAP = {
    "BROAD INDUSTRY": "broad_industry",
    "INDUSTRY SUBTYPE": "industry_subtype",
    "SPECIFIC ACTIVITY / PROCESS": "activity",
    "JURISDICTION LEVEL": "jurisdiction_level",
    "GOVERNING AGENCY": "governing_agency",
    "REGULATION / RULE NAME": "regulation_name",
    "CFR / USC CITATION": "citation",
    "REPORT / FORM NAME": "report_form_name",
    "FORM # / CODE": "form_code",
    "FILING FREQUENCY": "filing_frequency",
    "KEY DUE DATES / DEADLINES": "due_dates",
    "BUSINESS SEGMENT": "business_segment",
    "PENALTIES / CONSEQUENCES": "penalties",
    "CIP": "cip",
    "SIC": "sic",
    "NAICS": "naics",
    "SOC": "soc",
    "ISIC": "isic",
    "HS/HTS": "hs_hts",
    "NOTES": "notes",
    "UCO_NODE_ID": "uco_node_id",
    "ONTOLOGY_LEVEL": "ontology_level",
    "COMPLIANCE_CHAIN_REF": "compliance_chain_ref",
    "OPERATING_SEGMENT": "operating_segment",
    "RESPONSIBLE_ROLE": "responsible_role",
    "ENFORCEMENT_TYPE": "enforcement_type",
    "RISK_WEIGHT": "risk_weight",
    "YBR_GATE": "ybr_gate",
    "POLICY_ACTION": "policy_action",
    "LAST_UPDATED": "last_updated",
}

# Sheets that are reference material, not node sources
SKIP_SHEETS = {
    "INDEX", "CODE CROSSWALK", "LEGEND & FILING CALENDAR", "AGENCY REGISTRY",
    "NAICS FULL DECODER", "SIC DIVISION CROSSWALK", "SOC OCCUPATION CROSSWALK",
}

# Explicit per-node corrections (auditable). Each entry is logged in the
# validation report.
NODE_PATCHES = {
    # Column shift in source row: SIC value duplicated into NAICS slot,
    # NAICS->SOC, SOC->ISIC; ISIC lost. Chain ref also corrupted.
    "UCO-ENR-1037": {
        "fields": {
            "cip": "15.0301",
            "sic": "4911",
            "naics": "221112",
            "soc": "51-8013",
            "isic": "D35",
            "hs_hts": "2716",
        },
        "rebuild_chain": True,
        "reason": "Source row column shift (SIC in NAICS slot, cascading); "
                  "crosswalk realigned, compliance chain rebuilt.",
    },
}

VALID_POLICY_ACTIONS = {"BLOCK", "ESCALATE", "APPROVE"}
VALID_ENFORCEMENT = {
    "Civil Monetary Penalty", "Criminal", "Administrative",
    "License/Certificate", "Injunctive", "Warning/Notice",
}

NODE_ID_RE = re.compile(r"^UCO-[A-Z]{3}-\d{4}$")
CFR_RE = re.compile(r"(\d+)\s*CFR\s*(?:Parts?\s*)?(\d+)", re.IGNORECASE)


def sector_partition(sheet_name: str) -> str:
    """'01 – ENERGY' -> 'energy'; 'CROSS-CUTTING REGS' -> 'xsc' (matches V3)."""
    if sheet_name.upper().startswith("CROSS-CUTTING"):
        return "xsc"
    name = re.sub(r"^\d+\s*[–-]\s*", "", sheet_name)
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s != "—" else None


def parse_sheet(ws, sheet_name, report):
    rows = list(ws.iter_rows(values_only=True))
    hdr_i, idx = None, {}
    for i, r in enumerate(rows[:6]):
        if r and any(c == "UCO_NODE_ID" for c in r if c):
            hdr_i = i
            idx = {clean(c): j for j, c in enumerate(r) if clean(c)}
            break
    if hdr_i is None:
        report["skipped_sheets"].append(sheet_name)
        return []

    part = sector_partition(sheet_name)
    nodes = []
    for rn, r in enumerate(rows[hdr_i + 1:], start=hdr_i + 2):
        if not r or not any(r):
            continue
        rec = {}
        for header, field in HEADER_MAP.items():
            j = idx.get(header)
            rec[field] = clean(r[j]) if j is not None and j < len(r) else None
        if not rec.get("uco_node_id"):
            continue
        rec["sector_partition"] = part
        rec["_source"] = f"{sheet_name}!row{rn}"
        nodes.append(rec)
    return nodes


def rebuild_chain(rec):
    return (f"CIP:{rec.get('cip')} → NAICS:{rec.get('naics')} → "
            f"{rec.get('jurisdiction_level')} → {rec.get('governing_agency')} → "
            f"{(rec.get('regulation_name') or '')[:60]}")


def apply_patches(nodes, report):
    by_id = {n["uco_node_id"]: n for n in nodes}
    for node_id, patch in NODE_PATCHES.items():
        n = by_id.get(node_id)
        if not n:
            report["warnings"].append(f"Patch target {node_id} not found in workbook")
            continue
        n.update(patch["fields"])
        if patch.get("rebuild_chain"):
            n["compliance_chain_ref"] = rebuild_chain(n)
        report["patched"].append(f"{node_id}: {patch['reason']}")


def validate(nodes, report):
    seen = Counter(n["uco_node_id"] for n in nodes)
    dupes = {k for k, c in seen.items() if c > 1}
    valid = []
    for n in nodes:
        nid, errs = n["uco_node_id"], []
        if nid in dupes:
            errs.append("duplicate uco_node_id")
        if not NODE_ID_RE.match(nid):
            errs.append(f"node id format: {nid}")
        if n.get("policy_action") not in VALID_POLICY_ACTIONS:
            errs.append(f"policy_action: {n.get('policy_action')!r}")
        try:
            rw = int(n.get("risk_weight"))
            if not 1 <= rw <= 10:
                errs.append(f"risk_weight out of range: {rw}")
            n["risk_weight"] = rw
        except (TypeError, ValueError):
            errs.append(f"risk_weight not int: {n.get('risk_weight')!r}")
        if n.get("enforcement_type") and n["enforcement_type"] not in VALID_ENFORCEMENT:
            report["warnings"].append(
                f"{nid}: unrecognized enforcement_type {n['enforcement_type']!r}")
        naics = n.get("naics")
        if naics and not re.match(r"^[\d\-, ]+$", naics):
            # Cross-cutting obligations legitimately apply to ALL sectors.
            if n["sector_partition"] == "xsc" and naics.lower().startswith("all"):
                pass
            else:
                report["warnings"].append(f"{nid}: NAICS not numeric: {naics!r}")
        # last_updated -> ISO date or null
        lu = n.get("last_updated")
        if lu:
            m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(lu))
            n["last_updated"] = m.group(0) if m else None
            if not m:
                report["warnings"].append(f"{nid}: unparseable last_updated {lu!r}")
        if errs:
            report["rejected"].append(f"{nid} ({n['_source']}): " + "; ".join(errs))
        else:
            valid.append(n)
    return valid


def sql_quote(v):
    if v is None:
        return "NULL"
    if isinstance(v, int):
        return str(v)
    return "'" + str(v).replace("'", "''") + "'"


def emit_sql(nodes, path: Path):
    cols = SCHEMA_MAP["columns"]
    table, key = SCHEMA_MAP["table"], SCHEMA_MAP["conflict_key"]
    col_list = ", ".join(cols.values())
    update_cols = [c for f, c in cols.items() if c != key]
    update_set = ",\n      ".join(f"{c} = EXCLUDED.{c}" for c in update_cols)
    lines = [
        "-- R__seed_uco_nodes.sql",
        "-- GENERATED by convert_matrix_seed.py — do not edit by hand.",
        f"-- Source: SMEPro COS+ Universal Compliance Decoding Matrix",
        f"-- Generated: {date.today().isoformat()} | nodes: {len(nodes)}",
        "-- Repeatable migration: Flyway re-applies on checksum change (matrix updates).",
        "-- ALIGN column names with V4/V6 DDL via SCHEMA_MAP before first run.",
        "",
        "BEGIN;",
        "",
    ]
    for n in sorted(nodes, key=lambda x: x["uco_node_id"]):
        vals = ", ".join(sql_quote(n.get(f)) for f in cols)
        lines.append(
            f"INSERT INTO {table} ({col_list})\nVALUES ({vals})\n"
            f"ON CONFLICT ({key}) DO UPDATE SET\n      {update_set};\n"
        )
    lines += ["COMMIT;", ""]
    path.write_text("\n".join(lines), encoding="utf-8")


def ecfr_url(citation):
    """Best-effort eCFR part-level URL from a citation string."""
    m = CFR_RE.search(citation or "")
    if not m:
        return None
    title, part = m.group(1), m.group(2)
    return f"https://www.ecfr.gov/current/title-{title}/part-{part}"


def monitor_schedule(risk_weight):
    if risk_weight >= 10:
        return "every 12 hours"
    if risk_weight >= 9:
        return "every 24 hours"
    return "every 3 days"


def emit_monitors(nodes, path: Path, report):
    payloads, skipped = [], 0
    for n in nodes:
        if n["risk_weight"] < 8:
            continue
        if not (n.get("jurisdiction_level") or "").startswith("Federal"):
            continue
        url = ecfr_url(n.get("citation"))
        if not url:
            skipped += 1
            continue
        payloads.append({
            "uco_node_id": n["uco_node_id"],
            "tool": "firecrawl_monitor_create",
            "arguments": {
                "page": url,
                "metadata": {"uco_node_id": n["uco_node_id"]},
                "name": f"{n['uco_node_id']} — {(n.get('regulation_name') or '')[:60]}",
                "goal": (
                    f"Alert when the regulatory text, penalty amounts, filing "
                    f"deadlines, reporting thresholds, or form requirements under "
                    f"{n.get('citation')} ({n.get('regulation_name')}) change. "
                    f"Ignore formatting, navigation, or styling changes."
                ),
                "schedule": monitor_schedule(n["risk_weight"]),
            },
        })
    path.write_text(json.dumps(payloads, indent=2), encoding="utf-8")
    report["monitors"] = (
        f"{len(payloads)} payloads emitted (risk_weight>=8, Federal, CFR-citable); "
        f"{skipped} high-risk federal nodes skipped (no parseable CFR citation — "
        f"USC/state citations need manual page selection)."
    )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--out-dir", default="seed")
    ap.add_argument("--emit-monitors", action="store_true")
    args = ap.parse_args()

    out = Path(args.out_dir)
    out.mkdir(parents=True, exist_ok=True)
    report = {"skipped_sheets": [], "patched": [], "warnings": [],
              "rejected": [], "monitors": None}

    wb = load_workbook(args.xlsx, read_only=True)
    nodes, per_sheet = [], {}
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        parsed = parse_sheet(wb[name], name, report)
        if parsed:
            per_sheet[name] = len(parsed)
            nodes.extend(parsed)

    apply_patches(nodes, report)
    valid = validate(nodes, report)
    for n in valid:
        n.pop("_source", None)

    (out / "uco_nodes_seed.json").write_text(
        json.dumps(valid, indent=2), encoding="utf-8")
    emit_sql(valid, out / "R__seed_uco_nodes.sql")
    if args.emit_monitors:
        emit_monitors(valid, out / "firecrawl_monitors.json", report)

    actions = Counter(n["policy_action"] for n in valid)
    parts = Counter(n["sector_partition"] for n in valid)
    rpt = [
        "# Seed validation report", "",
        f"Generated: {date.today().isoformat()}",
        f"Source workbook: `{Path(args.xlsx).name}`", "",
        f"**Nodes parsed:** {len(nodes)}  |  **valid:** {len(valid)}  |  "
        f"**rejected:** {len(report['rejected'])}", "",
        "## Per-sheet counts", "",
        *[f"- {s}: {c}" for s, c in per_sheet.items()], "",
        "## Policy action distribution", "",
        *[f"- {k}: {v}" for k, v in sorted(actions.items())], "",
        f"## Partitions ({len(parts)})", "",
        *[f"- {k}: {v}" for k, v in sorted(parts.items())], "",
        "## Patches applied", "",
        *([f"- {p}" for p in report["patched"]] or ["- none"]), "",
        "## Rejected rows", "",
        *([f"- {r}" for r in report["rejected"]] or ["- none"]), "",
        "## Warnings", "",
        *([f"- {w}" for w in report["warnings"]] or ["- none"]), "",
    ]
    if report["monitors"]:
        rpt += ["## Firecrawl monitors", "", f"- {report['monitors']}", ""]
    (out / "seed_validation_report.md").write_text("\n".join(rpt), encoding="utf-8")

    print(f"valid={len(valid)} rejected={len(report['rejected'])} "
          f"warnings={len(report['warnings'])} -> {out}/")
    if report["rejected"]:
        sys.exit(2)


if __name__ == "__main__":
    main()
