#!/usr/bin/env python3
"""
load_uco_seeds.py
IOS+ UCO Matrix Seed Loader
EB Doc 6 §3.3

Loads the Universal Compliance Decoding Matrix into COS+ database.

This script supports two input modes:
  1. CSV mode (recommended): Reads pre-processed CSV files from preprocess_workbook.py
  2. XLSX mode (legacy): Reads directly from the source workbook (positional fallback)

Dependency order:
  1. agency_registry          — must exist before uco_nodes (foreign agency refs)
  2. naics_decoder            — must exist before uco_nodes
  3. uco_nodes                — core 350-node matrix
  4. code_crosswalk           — CIP/SIC/NAICS/SOC/ISIC/HS-HTS mappings
  5. uco_obligation_metadata  — provenance and trust metadata (V8 migration)
  6. compliance_chains        — chain definitions referencing uco_node_ids

Usage (CSV mode — recommended):
  python3 scripts/db/preprocess_workbook.py \\
      --xlsx SMEPro_COS_Universal_Compliance_Decoding_Matrix.xlsx \\
      --output-dir db/seeds/

  python3 scripts/db/load_uco_seeds.py \\
      --csv-dir db/seeds/ \\
      --db-url $DATABASE_URL_COS_ADMIN

Usage (XLSX mode — legacy):
  python3 scripts/db/load_uco_seeds.py \\
      --xlsx SMEPro_COS_Universal_Compliance_Decoding_Matrix.xlsx \\
      --db-url $DATABASE_URL_COS_ADMIN
"""

import argparse
import csv
import sys
import os
import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from datetime import date

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# Column mapping: 30 columns per UCO node (positional, used in XLSX legacy mode)
REGULATORY_COLS = [
    "broad_industry", "industry_subtype", "specific_activity", "jurisdiction_level",
    "governing_agency", "regulation_name", "cfr_usc_citation", "report_form_name",
    "form_code", "filing_frequency", "key_due_dates", "business_segment",
    "penalties_consequences", "cip", "sic", "naics", "soc", "isic", "hs_hts", "notes"
]
ENGINE_COLS = [
    "uco_node_id", "ontology_level", "compliance_chain_ref", "operating_segment",
    "responsible_role", "enforcement_type", "risk_weight", "ybr_gate",
    "policy_action", "last_updated"
]
ALL_COLS = REGULATORY_COLS + ENGINE_COLS

UCO_NODES_INSERT_COLS = ALL_COLS  # excludes internal _prefixed keys

AGENCY_REGISTRY_COLS = [
    "agency_code", "agency_name", "jurisdiction", "parent_agency", "website", "notes",
]

NAICS_DECODER_COLS = [
    "naics_code", "description", "sector_code", "sector_name",
    "subsector", "industry_grp", "naics_year",
]

CODE_CROSSWALK_COLS = [
    "code_system", "source_code", "target_system", "target_code", "confidence", "notes",
]

OBLIGATION_METADATA_COLS = [
    "uco_node_id", "report_family", "jurisdiction_detail", "state",
    "input_source", "submission_channel", "renderer_ref", "obligation_schema_id",
    "as_of_date", "verification_status", "source_note",
]

SKIP_SHEETS = {"INDEX", "AGENCY REGISTRY", "CODE CROSSWALK", "NAICS FULL DECODER"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean(val):
    """Strip and normalize empty/null string values to None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in {"none", "n/a", "na", "#n/a"} else None


def read_csv_records(path: str) -> list[dict]:
    """Read a CSV file and return a list of dicts. Raises FileNotFoundError if absent."""
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------------------
# CSV-mode loaders (header-driven, production path)
# ---------------------------------------------------------------------------

def load_agency_registry_from_csv(csv_path: str, cur, conn) -> int:
    """Load agency_registry from pre-processed CSV."""
    records = read_csv_records(csv_path)
    if not records:
        print(f"  agency_registry: 0 rows (empty CSV)")
        return 0

    sql = """
        INSERT INTO agency_registry (agency_code, agency_name, jurisdiction,
                                     parent_agency, website, notes)
        VALUES %s
        ON CONFLICT (agency_code) DO UPDATE SET
          agency_name  = EXCLUDED.agency_name,
          jurisdiction = EXCLUDED.jurisdiction,
          parent_agency = EXCLUDED.parent_agency,
          website      = EXCLUDED.website,
          notes        = EXCLUDED.notes
    """
    values = [
        (
            clean(r.get("agency_code")),
            clean(r.get("agency_name")),
            clean(r.get("jurisdiction")),
            clean(r.get("parent_agency")),
            clean(r.get("website")),
            clean(r.get("notes")),
        )
        for r in records
        if clean(r.get("agency_code"))
    ]
    execute_values(cur, sql, values)
    conn.commit()
    print(f"  agency_registry: {len(values)} rows upserted")
    return len(values)


def load_naics_decoder_from_csv(csv_path: str, cur, conn) -> int:
    """Load naics_decoder from pre-processed CSV."""
    records = read_csv_records(csv_path)
    if not records:
        print(f"  naics_decoder: 0 rows (empty CSV)")
        return 0

    sql = """
        INSERT INTO naics_decoder (naics_code, description, sector_code, sector_name,
                                   subsector, industry_grp, naics_year)
        VALUES %s
        ON CONFLICT (naics_code) DO UPDATE SET
          description  = EXCLUDED.description,
          sector_code  = EXCLUDED.sector_code,
          sector_name  = EXCLUDED.sector_name,
          subsector    = EXCLUDED.subsector,
          industry_grp = EXCLUDED.industry_grp,
          naics_year   = EXCLUDED.naics_year
    """
    values = []
    for r in records:
        naics_code = clean(r.get("naics_code"))
        if not naics_code:
            continue
        try:
            naics_year = int(r.get("naics_year") or 2022)
        except (TypeError, ValueError):
            naics_year = 2022
        values.append((
            naics_code,
            clean(r.get("description")),
            clean(r.get("sector_code")),
            clean(r.get("sector_name")),
            clean(r.get("subsector")),
            clean(r.get("industry_grp")),
            naics_year,
        ))
    execute_values(cur, sql, values)
    conn.commit()
    print(f"  naics_decoder: {len(values)} rows upserted")
    return len(values)


def load_uco_nodes_from_csv(csv_path: str, cur, conn) -> int:
    """Load uco_nodes from pre-processed CSV (header-driven)."""
    records = read_csv_records(csv_path)
    if not records:
        print("  uco_nodes: 0 rows (empty CSV)")
        return 0

    nodes = []
    for r in records:
        uco_id = clean(r.get("uco_node_id"))
        if not uco_id:
            continue
        try:
            risk_weight = int(r.get("risk_weight") or 5)
        except (TypeError, ValueError):
            risk_weight = 5

        last_updated_raw = clean(r.get("last_updated"))
        last_updated = last_updated_raw if last_updated_raw else str(date.today())

        nodes.append(tuple(
            risk_weight if col == "risk_weight"
            else last_updated if col == "last_updated"
            else clean(r.get(col))
            for col in UCO_NODES_INSERT_COLS
        ))

    INSERT_SQL = f"""
        INSERT INTO uco_nodes ({",".join(UCO_NODES_INSERT_COLS)})
        VALUES %s
        ON CONFLICT (uco_node_id) DO UPDATE SET
          regulation_name = EXCLUDED.regulation_name,
          risk_weight     = EXCLUDED.risk_weight,
          policy_action   = EXCLUDED.policy_action,
          ontology_level  = EXCLUDED.ontology_level,
          ybr_gate        = EXCLUDED.ybr_gate,
          jurisdiction_level = EXCLUDED.jurisdiction_level,
          last_updated    = EXCLUDED.last_updated
    """
    execute_values(cur, INSERT_SQL, nodes)
    conn.commit()
    print(f"  uco_nodes: {len(nodes)} rows upserted (expected 350)")
    return len(nodes)


def load_code_crosswalk_from_csv(csv_path: str, cur, conn) -> int:
    """Load code_crosswalk from pre-processed CSV."""
    records = read_csv_records(csv_path)
    if not records:
        print("  code_crosswalk: 0 rows (empty CSV)")
        return 0

    sql = """
        INSERT INTO code_crosswalk (code_system, source_code, target_system,
                                    target_code, confidence, notes)
        VALUES %s
        ON CONFLICT DO NOTHING
    """
    values = []
    for r in records:
        cs = clean(r.get("code_system"))
        sc = clean(r.get("source_code"))
        ts = clean(r.get("target_system"))
        tc = clean(r.get("target_code"))
        if not all([cs, sc, ts, tc]):
            continue
        try:
            confidence = float(r.get("confidence") or 1.0)
        except (TypeError, ValueError):
            confidence = 1.0
        values.append((cs, sc, ts, tc, round(confidence, 3), clean(r.get("notes"))))
    execute_values(cur, sql, values)
    conn.commit()
    print(f"  code_crosswalk: {len(values)} rows inserted (skip conflicts)")
    return len(values)


def load_obligation_metadata_from_csv(csv_path: str, cur, conn) -> int:
    """
    Load uco_obligation_metadata from pre-processed CSV.
    Requires V8 migration to have been applied.
    """
    records = read_csv_records(csv_path)
    if not records:
        print("  uco_obligation_metadata: 0 rows (empty CSV)")
        return 0

    sql = """
        INSERT INTO uco_obligation_metadata (
            uco_node_id, report_family, jurisdiction_detail, state,
            input_source, submission_channel, renderer_ref, obligation_schema_id,
            as_of_date, verification_status, source_note
        )
        VALUES %s
        ON CONFLICT (uco_node_id) DO UPDATE SET
          report_family        = EXCLUDED.report_family,
          jurisdiction_detail  = EXCLUDED.jurisdiction_detail,
          state                = EXCLUDED.state,
          input_source         = EXCLUDED.input_source,
          submission_channel   = EXCLUDED.submission_channel,
          renderer_ref         = EXCLUDED.renderer_ref,
          obligation_schema_id = EXCLUDED.obligation_schema_id,
          as_of_date           = EXCLUDED.as_of_date,
          verification_status  = EXCLUDED.verification_status,
          source_note          = EXCLUDED.source_note,
          last_verified_at     = CASE
            WHEN EXCLUDED.verification_status = 'verified' THEN now()
            ELSE uco_obligation_metadata.last_verified_at
          END
    """
    values = []
    for r in records:
        uco_id = clean(r.get("uco_node_id"))
        if not uco_id:
            continue
        vstatus = clean(r.get("verification_status")) or "pending"
        as_of = clean(r.get("as_of_date")) or None
        values.append((
            uco_id,
            clean(r.get("report_family")),
            clean(r.get("jurisdiction_detail")),
            clean(r.get("state")),
            clean(r.get("input_source")),
            clean(r.get("submission_channel")),
            clean(r.get("renderer_ref")),
            clean(r.get("obligation_schema_id")),
            as_of,
            vstatus,
            clean(r.get("source_note")),
        ))
    execute_values(cur, sql, values)
    conn.commit()
    print(f"  uco_obligation_metadata: {len(values)} rows upserted")
    return len(values)


# ---------------------------------------------------------------------------
# XLSX legacy mode (positional, unchanged from original)
# ---------------------------------------------------------------------------

def load_nodes_from_xlsx(xlsx_path: str, db_url: str):
    """
    Legacy XLSX loader. Reads uco_nodes positionally from sector sheets.
    Use preprocess_workbook.py + CSV mode for production ingestion.
    """
    print(f"Loading UCO matrix from: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    conn = psycopg2.connect(db_url)
    cur = conn.cursor()

    nodes = []
    seen_ids = set()
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS or not sheet_name[0].isdigit() and "CROSS" not in sheet_name:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if len(row) < 30 or not row[20]:
                continue
            uco_id = clean(row[20])
            if not uco_id or uco_id in seen_ids:
                continue
            seen_ids.add(uco_id)
            node = {col: clean(row[i]) for i, col in enumerate(ALL_COLS)}
            # Coerce types
            node["risk_weight"] = int(node["risk_weight"] or 5)
            node["last_updated"] = date.today()
            nodes.append(node)

    print(f"Nodes parsed: {len(nodes)} (expected 350)")

    INSERT_SQL = f"""
        INSERT INTO uco_nodes ({",".join(ALL_COLS)})
        VALUES %s
        ON CONFLICT (uco_node_id) DO UPDATE SET
          regulation_name = EXCLUDED.regulation_name,
          risk_weight = EXCLUDED.risk_weight,
          policy_action = EXCLUDED.policy_action,
          last_updated = EXCLUDED.last_updated
    """
    values = [tuple(n[c] for c in ALL_COLS) for n in nodes]
    execute_values(cur, INSERT_SQL, values)
    conn.commit()
    conn.close()
    print(f"Loaded {len(nodes)} UCO nodes into COS+")


# ---------------------------------------------------------------------------
# CSV-mode pipeline orchestrator
# ---------------------------------------------------------------------------

def load_from_csv_dir(csv_dir: str, db_url: str, skip_metadata: bool = False) -> None:
    """
    Load all seed tables from a directory of pre-processed CSVs.
    Expected files (produced by preprocess_workbook.py):
      agency_registry.csv, naics_decoder.csv, uco_nodes.csv,
      code_crosswalk.csv, obligation_metadata.csv
    """
    print(f"Loading seed CSVs from: {csv_dir}")
    conn = psycopg2.connect(db_url)
    cur = conn.cursor()

    total = 0

    # 1. agency_registry
    path = os.path.join(csv_dir, "agency_registry.csv")
    if os.path.isfile(path):
        total += load_agency_registry_from_csv(path, cur, conn)
    else:
        print(f"  agency_registry.csv not found; skipping.")

    # 2. naics_decoder
    path = os.path.join(csv_dir, "naics_decoder.csv")
    if os.path.isfile(path):
        total += load_naics_decoder_from_csv(path, cur, conn)
    else:
        print(f"  naics_decoder.csv not found; skipping.")

    # 3. uco_nodes
    path = os.path.join(csv_dir, "uco_nodes.csv")
    if os.path.isfile(path):
        total += load_uco_nodes_from_csv(path, cur, conn)
    else:
        print(f"ERROR: uco_nodes.csv not found in {csv_dir}", file=sys.stderr)
        cur.close()
        conn.close()
        sys.exit(1)

    # 4. code_crosswalk
    path = os.path.join(csv_dir, "code_crosswalk.csv")
    if os.path.isfile(path):
        total += load_code_crosswalk_from_csv(path, cur, conn)
    else:
        print(f"  code_crosswalk.csv not found; skipping.")

    # 5. obligation_metadata (requires V8 migration)
    if not skip_metadata:
        path = os.path.join(csv_dir, "obligation_metadata.csv")
        if os.path.isfile(path):
            try:
                total += load_obligation_metadata_from_csv(path, cur, conn)
            except psycopg2.errors.UndefinedTable:
                conn.rollback()
                print(
                    "WARNING: uco_obligation_metadata table not found. "
                    "Apply V8 migration first, or use --skip-metadata.",
                    file=sys.stderr,
                )
        else:
            print(f"  obligation_metadata.csv not found; skipping.")

    cur.close()
    conn.close()
    print(f"\nSeed load complete. Total rows upserted/inserted: {total}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="IOS+ UCO Matrix Seed Loader",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument(
        "--xlsx",
        metavar="PATH",
        help="(Legacy) Source .xlsx workbook. Loads only uco_nodes positionally.",
    )
    input_group.add_argument(
        "--csv-dir",
        metavar="DIR",
        help=(
            "(Recommended) Directory of pre-processed CSVs from preprocess_workbook.py. "
            "Loads all seed tables including uco_obligation_metadata."
        ),
    )

    parser.add_argument(
        "--db-url",
        default=os.environ.get("DATABASE_URL_COS_ADMIN"),
        metavar="URL",
        help="PostgreSQL connection URL (default: $DATABASE_URL_COS_ADMIN)",
    )
    parser.add_argument(
        "--skip-metadata",
        action="store_true",
        default=False,
        help="Skip loading uco_obligation_metadata (if V8 migration not yet applied)",
    )
    args = parser.parse_args()

    if not args.db_url:
        print("ERROR: --db-url or DATABASE_URL_COS_ADMIN required", file=sys.stderr)
        sys.exit(1)

    if args.xlsx:
        load_nodes_from_xlsx(args.xlsx, args.db_url)
    else:
        if not os.path.isdir(args.csv_dir):
            print(f"ERROR: --csv-dir not found: {args.csv_dir}", file=sys.stderr)
            sys.exit(1)
        load_from_csv_dir(args.csv_dir, args.db_url, skip_metadata=args.skip_metadata)


if __name__ == "__main__":
    main()
