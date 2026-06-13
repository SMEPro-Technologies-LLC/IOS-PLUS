#!/usr/bin/env python3
# /// script
# dependencies = [
#   "openpyxl>=3.1.0",
# ]
# ///
"""
preprocess_workbook.py
IOS+ UCO Workbook Ingestion Preprocessor
EB Doc 6 §3.3 / UCO Obligation Metadata Pipeline

Reads the UCO workbook and transforms each logical sheet into
repo-compatible seed CSVs and a structured transformation report.

Supported sheets:
  - uco_nodes              → seeds/uco_nodes.csv
  - agency_registry        → seeds/agency_registry.csv
  - naics_decoder          → seeds/naics_decoder.csv
  - code_crosswalk         → seeds/code_crosswalk.csv
  - _obligation_metadata   → seeds/obligation_metadata.csv

Normalization applied:
  - ontology_level: maps raw workbook labels to valid DB values
    (sector | subsector | activity | cross-cutting)
  - ybr_gate: maps "Gate 530: Compliance Check" → "L5", etc.
  - jurisdiction_level: maps "State – TX", "Federal / State" →
    canonical DB value, preserving detail in obligation_metadata

Emits:
  - <output_dir>/uco_nodes.csv
  - <output_dir>/agency_registry.csv
  - <output_dir>/naics_decoder.csv
  - <output_dir>/code_crosswalk.csv
  - <output_dir>/obligation_metadata.csv
  - <output_dir>/transform_report.json

Usage:
  python3 scripts/db/preprocess_workbook.py \\
      --xlsx SMEPro_COS_Universal_Compliance_Decoding_Matrix.xlsx \\
      --output-dir db/seeds/

Exit codes:
  0  — success
  1  — invalid arguments or missing required environment
  2  — workbook structure validation failure
  3  — normalization produced zero valid rows
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import date, datetime
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover
    print("FATAL: openpyxl not found. Install openpyxl>=3.1.0.", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Valid DB constraint values
# ---------------------------------------------------------------------------

VALID_ONTOLOGY_LEVELS = frozenset({"sector", "subsector", "activity", "cross-cutting"})
VALID_YBR_GATES = frozenset({"L3", "L4", "L5", "L7"})
VALID_JURISDICTION_LEVELS = frozenset({"Federal", "State", "Local", "International"})
VALID_VERIFICATION_STATUSES = frozenset({"verified", "stale", "corrected", "pending"})

# ---------------------------------------------------------------------------
# Normalization maps
# ---------------------------------------------------------------------------

# ontology_level raw → canonical DB value
ONTOLOGY_LEVEL_MAP: dict[str, str] = {
    # exact canonical forms
    "sector": "sector",
    "subsector": "subsector",
    "activity": "activity",
    "cross-cutting": "cross-cutting",
    "crosscutting": "cross-cutting",
    # common workbook labels
    "l2: regulations & rules": "subsector",
    "l2:regulations & rules": "subsector",
    "regulations & rules": "subsector",
    "regulations and rules": "subsector",
    "l1: broad industry": "sector",
    "l1:broad industry": "sector",
    "broad industry": "sector",
    "l3: specific activity": "activity",
    "l3:specific activity": "activity",
    "specific activity": "activity",
    "functional": "activity",  # legacy label
    "cross cutting": "cross-cutting",
    "xsc": "cross-cutting",
    "xsc-cross-cutting": "cross-cutting",
    # edge cases observed in workbooks
    "sub-sector": "subsector",
    "sub sector": "subsector",
    "industry subtype": "subsector",
}

# ybr_gate raw → canonical DB value
YBR_GATE_MAP: dict[str, str] = {
    # canonical forms
    "l3": "L3",
    "l4": "L4",
    "l5": "L5",
    "l7": "L7",
    # descriptive labels
    "gate 530: compliance check": "L5",
    "gate530: compliance check": "L5",
    "gate 530 compliance check": "L5",
    "gate530 compliance check": "L5",
    "gate 530": "L5",
    "gate530": "L5",
    "l5: gate 530": "L5",
    "l3: ontological mapping": "L3",
    "l3:ontological mapping": "L3",
    "l3 ontological mapping": "L3",
    "ontological mapping": "L3",
    "l4: evidence collection": "L4",
    "l4:evidence collection": "L4",
    "l4 evidence collection": "L4",
    "evidence collection": "L4",
    "l7: synthesis": "L7",
    "l7:synthesis": "L7",
    "l7 synthesis": "L7",
    "synthesis": "L7",
}

# Raw jurisdiction → (canonical_db_value, state_code_or_None)
# The canonical DB value must be one of VALID_JURISDICTION_LEVELS.
# State code is extracted when applicable.
def _parse_jurisdiction(raw: str | None) -> tuple[str, str | None]:
    """
    Normalize a raw jurisdiction string to (canonical_level, state_code).

    Returns:
        (jurisdiction_level, state)
        where jurisdiction_level ∈ VALID_JURISDICTION_LEVELS
        and state is a 2-letter state code or None.

    Raises ValueError if the raw value cannot be mapped.
    """
    if not raw:
        raise ValueError("jurisdiction_level is required but empty")

    normalized = raw.strip()
    lower = normalized.lower()

    # Fast-path exact canonical matches
    for valid in VALID_JURISDICTION_LEVELS:
        if normalized == valid:
            return (valid, None)

    # "State – TX" pattern
    state_dash = re.match(
        r"^state\s*[–—\-]\s*([A-Z]{2})\b",
        normalized,
        re.IGNORECASE,
    )
    if state_dash:
        return ("State", state_dash.group(1).upper())

    # "State (TX)" pattern
    state_paren = re.match(r"^state\s*\(\s*([A-Z]{2})\s*\)", normalized, re.IGNORECASE)
    if state_paren:
        return ("State", state_paren.group(1).upper())

    # "Federal / State" or "Federal/State" → resolve to Federal (broader jurisdiction)
    if re.match(r"^federal\s*/\s*state$", lower):
        return ("Federal", None)

    # "Federal / Local" → Federal
    if re.match(r"^federal\s*/\s*local$", lower):
        return ("Federal", None)

    # "State / Local" → State
    if re.match(r"^state\s*/\s*local$", lower):
        return ("State", None)

    # Prefix-based fallbacks
    if lower.startswith("state"):
        return ("State", None)
    if lower.startswith("federal"):
        return ("Federal", None)
    if lower.startswith("local"):
        return ("Local", None)
    if lower.startswith("international"):
        return ("International", None)

    raise ValueError(f"Cannot map jurisdiction to valid DB value: {raw!r}")


def normalize_ontology_level(raw: str | None) -> str:
    """Map a raw workbook ontology_level label to a valid DB value."""
    if not raw:
        raise ValueError("ontology_level is required but empty")
    key = raw.strip().lower()
    if key in ONTOLOGY_LEVEL_MAP:
        return ONTOLOGY_LEVEL_MAP[key]
    raise ValueError(f"Cannot map ontology_level to valid DB value: {raw!r}")


def normalize_ybr_gate(raw: str | None) -> str:
    """Map a raw workbook ybr_gate label to a valid DB value."""
    if not raw:
        raise ValueError("ybr_gate is required but empty")
    key = raw.strip().lower()
    if key in YBR_GATE_MAP:
        return YBR_GATE_MAP[key]
    raise ValueError(f"Cannot map ybr_gate to valid DB value: {raw!r}")


def normalize_jurisdiction(raw: str | None) -> tuple[str, str | None]:
    """Wrapper around _parse_jurisdiction with a consistent signature."""
    return _parse_jurisdiction(raw)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean(val: Any) -> str | None:
    """Strip and return None for empty/None/"none" values."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in {"none", "n/a", "na", "#n/a"} else None


def header_index(headers: list[str], name: str, required: bool = True) -> int:
    """
    Find the 0-based index of a column header (case-insensitive, stripped).
    Raises ValueError if required and not found.
    Returns -1 if optional and not found.
    """
    lower_name = name.strip().lower()
    for i, h in enumerate(headers):
        if h.strip().lower() == lower_name:
            return i
    if required:
        raise ValueError(f"Required column '{name}' not found in headers: {headers}")
    return -1


def read_sheet_with_headers(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
) -> tuple[list[str], list[list[Any]]]:
    """
    Read a worksheet and return (headers, data_rows).
    Row 1 is treated as the header row.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return ([], [])
    headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
    data_rows = [list(row) for row in all_rows[1:] if any(v is not None for v in row)]
    return (headers, data_rows)


# ---------------------------------------------------------------------------
# Sheet processors
# ---------------------------------------------------------------------------

# Required columns per sheet (header-driven)
UCO_NODES_REQUIRED_COLS = [
    "uco_node_id", "broad_industry", "industry_subtype", "specific_activity",
    "jurisdiction_level", "governing_agency", "regulation_name", "naics",
    "ontology_level", "enforcement_type", "risk_weight", "ybr_gate", "policy_action",
]

AGENCY_REGISTRY_REQUIRED_COLS = [
    "agency_code", "agency_name", "jurisdiction",
]

NAICS_DECODER_REQUIRED_COLS = [
    "naics_code", "description", "sector_code", "sector_name",
]

CODE_CROSSWALK_REQUIRED_COLS = [
    "code_system", "source_code", "target_system", "target_code",
]

OBLIGATION_METADATA_REQUIRED_COLS = [
    "uco_node_id",
]


def _validate_required_cols(
    headers: list[str],
    required: list[str],
    sheet_name: str,
) -> None:
    """Raise ValueError if any required columns are missing from headers."""
    lower_headers = {h.strip().lower() for h in headers}
    missing = [c for c in required if c.strip().lower() not in lower_headers]
    if missing:
        raise ValueError(
            f"Sheet '{sheet_name}' is missing required columns: {missing}. "
            f"Found headers: {headers}"
        )


def process_uco_nodes_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    report: dict,
) -> list[dict]:
    """
    Process the uco_nodes sheet into a list of dicts ready for CSV output.
    Applies normalization for ontology_level, ybr_gate, jurisdiction_level.
    """
    headers, rows = read_sheet_with_headers(ws)
    _validate_required_cols(headers, UCO_NODES_REQUIRED_COLS, "uco_nodes")

    processed: list[dict] = []
    errors: list[dict] = []
    warnings: list[dict] = []
    seen_ids: set[str] = set()

    for row_num, row in enumerate(rows, start=2):
        def get(col: str, required: bool = False) -> str | None:
            idx = header_index(headers, col, required=required)
            if idx == -1:
                return None
            return clean(row[idx] if idx < len(row) else None)

        uco_node_id = get("uco_node_id", required=False)
        if not uco_node_id:
            continue
        if uco_node_id in seen_ids:
            warnings.append({
                "row": row_num, "uco_node_id": uco_node_id,
                "field": "uco_node_id",
                "raw": uco_node_id, "normalized": uco_node_id,
                "note": "Duplicate uco_node_id; row skipped",
            })
            continue
        seen_ids.add(uco_node_id)

        record: dict[str, Any] = {"uco_node_id": uco_node_id}

        # --- ontology_level normalization ---
        raw_ontology = get("ontology_level")
        try:
            record["ontology_level"] = normalize_ontology_level(raw_ontology)
            if raw_ontology != record["ontology_level"]:
                warnings.append({
                    "row": row_num, "uco_node_id": uco_node_id,
                    "field": "ontology_level",
                    "raw": raw_ontology, "normalized": record["ontology_level"],
                })
        except ValueError as exc:
            errors.append({"row": row_num, "uco_node_id": uco_node_id, "error": str(exc)})
            continue

        # --- ybr_gate normalization ---
        raw_gate = get("ybr_gate")
        try:
            record["ybr_gate"] = normalize_ybr_gate(raw_gate)
            if raw_gate != record["ybr_gate"]:
                warnings.append({
                    "row": row_num, "uco_node_id": uco_node_id,
                    "field": "ybr_gate",
                    "raw": raw_gate, "normalized": record["ybr_gate"],
                })
        except ValueError as exc:
            errors.append({"row": row_num, "uco_node_id": uco_node_id, "error": str(exc)})
            continue

        # --- jurisdiction_level normalization ---
        raw_jur = get("jurisdiction_level")
        try:
            jur_level, state_code = normalize_jurisdiction(raw_jur)
            record["jurisdiction_level"] = jur_level
            record["_jurisdiction_detail"] = raw_jur   # preserved for metadata
            record["_state_code"] = state_code
            if raw_jur != jur_level:
                warnings.append({
                    "row": row_num, "uco_node_id": uco_node_id,
                    "field": "jurisdiction_level",
                    "raw": raw_jur, "normalized": jur_level, "state": state_code,
                })
        except ValueError as exc:
            errors.append({"row": row_num, "uco_node_id": uco_node_id, "error": str(exc)})
            continue

        # --- Remaining columns (passed through) ---
        for col in [
            "broad_industry", "industry_subtype", "specific_activity",
            "governing_agency", "regulation_name", "cfr_usc_citation",
            "report_form_name", "form_code", "filing_frequency", "key_due_dates",
            "business_segment", "penalties_consequences", "cip", "sic", "naics",
            "soc", "isic", "hs_hts", "notes", "compliance_chain_ref",
            "operating_segment", "responsible_role", "enforcement_type",
            "risk_weight", "policy_action", "last_updated",
        ]:
            if col not in record:
                record[col] = get(col)

        # Coerce types
        try:
            record["risk_weight"] = int(record["risk_weight"] or 5)
        except (TypeError, ValueError):
            record["risk_weight"] = 5

        if not record.get("last_updated"):
            record["last_updated"] = str(date.today())

        processed.append(record)

    report["uco_nodes"] = {
        "rows_processed": len(rows),
        "rows_accepted": len(processed),
        "rows_error": len(errors),
        "normalization_warnings": len(warnings),
        "errors": errors,
        "warnings": warnings,
    }
    return processed


def process_agency_registry_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    report: dict,
) -> list[dict]:
    """Process agency_registry sheet into normalized dicts."""
    headers, rows = read_sheet_with_headers(ws)
    _validate_required_cols(headers, AGENCY_REGISTRY_REQUIRED_COLS, "agency_registry")

    processed: list[dict] = []
    errors: list[dict] = []

    for row_num, row in enumerate(rows, start=2):
        def get(col: str) -> str | None:
            idx = header_index(headers, col, required=False)
            return clean(row[idx] if idx != -1 and idx < len(row) else None)

        agency_code = get("agency_code")
        if not agency_code:
            continue

        raw_jur = get("jurisdiction")
        try:
            jur_level, _ = normalize_jurisdiction(raw_jur)
        except ValueError as exc:
            errors.append({"row": row_num, "agency_code": agency_code, "error": str(exc)})
            continue

        processed.append({
            "agency_code": agency_code,
            "agency_name": get("agency_name"),
            "jurisdiction": jur_level,
            "parent_agency": get("parent_agency"),
            "website": get("website"),
            "notes": get("notes"),
        })

    report["agency_registry"] = {
        "rows_processed": len(rows),
        "rows_accepted": len(processed),
        "rows_error": len(errors),
        "errors": errors,
    }
    return processed


def process_naics_decoder_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    report: dict,
) -> list[dict]:
    """Process naics_decoder sheet into normalized dicts."""
    headers, rows = read_sheet_with_headers(ws)
    _validate_required_cols(headers, NAICS_DECODER_REQUIRED_COLS, "naics_decoder")

    processed: list[dict] = []
    errors: list[dict] = []

    for row_num, row in enumerate(rows, start=2):
        def get(col: str) -> str | None:
            idx = header_index(headers, col, required=False)
            return clean(row[idx] if idx != -1 and idx < len(row) else None)

        naics_code = get("naics_code")
        if not naics_code:
            continue

        naics_year_raw = get("naics_year")
        try:
            naics_year = int(naics_year_raw) if naics_year_raw else 2022
        except ValueError:
            naics_year = 2022

        processed.append({
            "naics_code": naics_code,
            "description": get("description"),
            "sector_code": get("sector_code"),
            "sector_name": get("sector_name"),
            "subsector": get("subsector"),
            "industry_grp": get("industry_grp"),
            "naics_year": naics_year,
        })

    report["naics_decoder"] = {
        "rows_processed": len(rows),
        "rows_accepted": len(processed),
        "rows_error": len(errors),
        "errors": errors,
    }
    return processed


def process_code_crosswalk_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    report: dict,
) -> list[dict]:
    """Process code_crosswalk sheet into normalized dicts."""
    headers, rows = read_sheet_with_headers(ws)
    _validate_required_cols(headers, CODE_CROSSWALK_REQUIRED_COLS, "code_crosswalk")

    processed: list[dict] = []
    errors: list[dict] = []

    valid_systems = {"CIP", "SIC", "NAICS", "SOC", "ISIC", "HS/HTS"}

    for row_num, row in enumerate(rows, start=2):
        def get(col: str) -> str | None:
            idx = header_index(headers, col, required=False)
            return clean(row[idx] if idx != -1 and idx < len(row) else None)

        code_system = get("code_system")
        source_code = get("source_code")
        target_system = get("target_system")
        target_code = get("target_code")

        if not all([code_system, source_code, target_system, target_code]):
            continue

        # Normalize code system aliases
        cs_upper = code_system.upper() if code_system else ""
        ts_upper = target_system.upper() if target_system else ""
        if cs_upper == "HS-HTS" or cs_upper == "HS/HTS":
            cs_upper = "HS/HTS"
        if ts_upper == "HS-HTS" or ts_upper == "HS/HTS":
            ts_upper = "HS/HTS"

        if cs_upper not in valid_systems:
            errors.append({
                "row": row_num,
                "error": f"Invalid code_system '{code_system}'. Valid: {sorted(valid_systems)}",
            })
            continue
        if ts_upper not in valid_systems:
            errors.append({
                "row": row_num,
                "error": f"Invalid target_system '{target_system}'. Valid: {sorted(valid_systems)}",
            })
            continue

        confidence_raw = get("confidence")
        try:
            confidence = float(confidence_raw) if confidence_raw else 1.0
            confidence = max(0.0, min(1.0, confidence))
        except ValueError:
            confidence = 1.0

        processed.append({
            "code_system": cs_upper,
            "source_code": source_code,
            "target_system": ts_upper,
            "target_code": target_code,
            "confidence": round(confidence, 3),
            "notes": get("notes"),
        })

    report["code_crosswalk"] = {
        "rows_processed": len(rows),
        "rows_accepted": len(processed),
        "rows_error": len(errors),
        "errors": errors,
    }
    return processed


def process_obligation_metadata_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    uco_nodes_data: list[dict],
    report: dict,
) -> list[dict]:
    """
    Process _obligation_metadata sheet.
    Also merges in jurisdiction_detail + state_code extracted from uco_nodes processing.
    """
    headers, rows = read_sheet_with_headers(ws)
    _validate_required_cols(headers, OBLIGATION_METADATA_REQUIRED_COLS, "_obligation_metadata")

    # Build lookup of jurisdiction details from uco_nodes processing
    jur_detail_by_id: dict[str, tuple[str | None, str | None]] = {
        n["uco_node_id"]: (n.get("_jurisdiction_detail"), n.get("_state_code"))
        for n in uco_nodes_data
    }

    processed: list[dict] = []
    errors: list[dict] = []

    for row_num, row in enumerate(rows, start=2):
        def get(col: str) -> str | None:
            idx = header_index(headers, col, required=False)
            return clean(row[idx] if idx != -1 and idx < len(row) else None)

        uco_node_id = get("uco_node_id")
        if not uco_node_id:
            continue

        raw_vstatus = get("verification_status") or "pending"
        if raw_vstatus.lower() in VALID_VERIFICATION_STATUSES:
            vstatus = raw_vstatus.lower()
        else:
            vstatus = "pending"
            errors.append({
                "row": row_num,
                "uco_node_id": uco_node_id,
                "warning": f"Unknown verification_status '{raw_vstatus}', defaulting to 'pending'",
            })

        # Prefer jurisdiction_detail from workbook column; fall back to uco_nodes extraction
        jur_detail_wb = get("jurisdiction_detail")
        node_jur_detail, node_state = jur_detail_by_id.get(uco_node_id, (None, None))
        jurisdiction_detail = jur_detail_wb or node_jur_detail
        state = get("state") or node_state

        # Normalize as_of_date
        as_of_raw = get("as_of_date")
        as_of_date: str | None = None
        if as_of_raw:
            try:
                # Accept various date formats
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y%m%d"):
                    try:
                        as_of_date = datetime.strptime(as_of_raw, fmt).date().isoformat()
                        break
                    except ValueError:
                        continue
                if not as_of_date:
                    errors.append({
                        "row": row_num,
                        "uco_node_id": uco_node_id,
                        "warning": f"Could not parse as_of_date '{as_of_raw}', skipping",
                    })
            except Exception:
                pass

        processed.append({
            "uco_node_id": uco_node_id,
            "report_family": get("report_family"),
            "jurisdiction_detail": jurisdiction_detail,
            "state": state,
            "input_source": get("input_source"),
            "submission_channel": get("submission_channel"),
            "renderer_ref": get("renderer_ref"),
            "obligation_schema_id": get("obligation_schema_id"),
            "as_of_date": as_of_date,
            "verification_status": vstatus,
            "source_note": get("source_note"),
        })

    report["obligation_metadata"] = {
        "rows_processed": len(rows),
        "rows_accepted": len(processed),
        "rows_error": len(errors),
        "errors": errors,
    }
    return processed


def _build_obligation_metadata_from_nodes(uco_nodes_data: list[dict]) -> list[dict]:
    """
    For uco_nodes that had jurisdiction_detail extracted during normalization
    but were not represented in the _obligation_metadata sheet, synthesize
    a minimal metadata record to preserve provenance.
    """
    records: list[dict] = []
    for node in uco_nodes_data:
        jur_detail = node.get("_jurisdiction_detail")
        state = node.get("_state_code")
        if jur_detail and jur_detail != node.get("jurisdiction_level"):
            records.append({
                "uco_node_id": node["uco_node_id"],
                "report_family": None,
                "jurisdiction_detail": jur_detail,
                "state": state,
                "input_source": None,
                "submission_channel": None,
                "renderer_ref": None,
                "obligation_schema_id": None,
                "as_of_date": None,
                "verification_status": "pending",
                "source_note": "Auto-generated from uco_nodes jurisdiction normalization",
            })
    return records


# ---------------------------------------------------------------------------
# CSV writers
# ---------------------------------------------------------------------------

def write_csv(path: str, records: list[dict], fieldnames: list[str]) -> int:
    """Write records to a CSV file, returning the number of rows written."""
    os.makedirs(os.path.dirname(path) if os.path.dirname(path) else ".", exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=fieldnames, extrasaction="ignore", quoting=csv.QUOTE_MINIMAL
        )
        writer.writeheader()
        writer.writerows(records)
    return len(records)


# Column order definitions (match DB schema column order)
UCO_NODES_CSV_COLS = [
    "broad_industry", "industry_subtype", "specific_activity", "jurisdiction_level",
    "governing_agency", "regulation_name", "cfr_usc_citation", "report_form_name",
    "form_code", "filing_frequency", "key_due_dates", "business_segment",
    "penalties_consequences", "cip", "sic", "naics", "soc", "isic", "hs_hts", "notes",
    "uco_node_id", "ontology_level", "compliance_chain_ref", "operating_segment",
    "responsible_role", "enforcement_type", "risk_weight", "ybr_gate",
    "policy_action", "last_updated",
]

AGENCY_REGISTRY_CSV_COLS = [
    "agency_code", "agency_name", "jurisdiction", "parent_agency", "website", "notes",
]

NAICS_DECODER_CSV_COLS = [
    "naics_code", "description", "sector_code", "sector_name",
    "subsector", "industry_grp", "naics_year",
]

CODE_CROSSWALK_CSV_COLS = [
    "code_system", "source_code", "target_system", "target_code", "confidence", "notes",
]

OBLIGATION_METADATA_CSV_COLS = [
    "uco_node_id", "report_family", "jurisdiction_detail", "state",
    "input_source", "submission_channel", "renderer_ref", "obligation_schema_id",
    "as_of_date", "verification_status", "source_note",
]


# ---------------------------------------------------------------------------
# Sheet name resolution
# ---------------------------------------------------------------------------

# Acceptable sheet name variants (case-insensitive prefix/exact matches)
SHEET_ALIASES: dict[str, list[str]] = {
    "uco_nodes": ["uco_nodes", "uco nodes", "ucomatrix", "uco matrix", "nodes"],
    "agency_registry": ["agency_registry", "agency registry", "agencies"],
    "naics_decoder": ["naics_decoder", "naics decoder", "naics full decoder", "naics"],
    "code_crosswalk": ["code_crosswalk", "code crosswalk", "crosswalk"],
    "obligation_metadata": [
        "_obligation_metadata", "obligation_metadata", "obligation metadata",
        "metadata", "_metadata",
    ],
}


def find_sheet(wb: "openpyxl.Workbook", logical_name: str) -> "openpyxl.worksheet.worksheet.Worksheet | None":
    """Find a worksheet by logical name using alias matching (case-insensitive)."""
    aliases = SHEET_ALIASES.get(logical_name, [logical_name])
    lower_names = {sn.lower(): sn for sn in wb.sheetnames}
    for alias in aliases:
        if alias.lower() in lower_names:
            return wb[lower_names[alias.lower()]]
    # Partial match fallback
    for alias in aliases:
        for sname_lower, sname in lower_names.items():
            if alias.lower() in sname_lower:
                return wb[sname]
    return None


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run_pipeline(
    xlsx_path: str,
    output_dir: str,
    strict: bool = False,
) -> dict:
    """
    Run the full workbook ingestion preprocessing pipeline.

    Args:
        xlsx_path:   Path to the source .xlsx workbook.
        output_dir:  Directory to write seed CSVs and transform_report.json.
        strict:      If True, exit 2 on any normalization error. If False, warn.

    Returns:
        The transformation report dict.

    Raises:
        SystemExit with appropriate exit codes on fatal errors.
    """
    print(f"[preprocess_workbook] Loading workbook: {xlsx_path}")
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except FileNotFoundError:
        print(f"ERROR: Workbook not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"ERROR: Cannot open workbook: {exc}", file=sys.stderr)
        sys.exit(1)

    print(f"[preprocess_workbook] Sheets found: {wb.sheetnames}")

    report: dict = {
        "source_workbook": os.path.basename(xlsx_path),
        "output_dir": output_dir,
        "processed_at": datetime.utcnow().isoformat() + "Z",
        "sheets_found": wb.sheetnames,
    }

    os.makedirs(output_dir, exist_ok=True)

    # ── 1. uco_nodes ──────────────────────────────────────────
    ws_nodes = find_sheet(wb, "uco_nodes")
    uco_nodes_data: list[dict] = []
    if ws_nodes is None:
        msg = "Sheet 'uco_nodes' (or alias) not found in workbook."
        if strict:
            print(f"ERROR: {msg}", file=sys.stderr)
            sys.exit(2)
        print(f"WARNING: {msg}")
        report["uco_nodes"] = {"error": msg}
    else:
        print(f"[preprocess_workbook] Processing sheet: {ws_nodes.title}")
        try:
            uco_nodes_data = process_uco_nodes_sheet(ws_nodes, report)
        except ValueError as exc:
            print(f"ERROR: uco_nodes sheet validation failed: {exc}", file=sys.stderr)
            sys.exit(2)

        if len(uco_nodes_data) == 0:
            print("ERROR: uco_nodes produced 0 valid rows.", file=sys.stderr)
            sys.exit(3)

        n = write_csv(
            os.path.join(output_dir, "uco_nodes.csv"),
            uco_nodes_data,
            UCO_NODES_CSV_COLS,
        )
        print(f"[preprocess_workbook] uco_nodes.csv: {n} rows")
        if strict and report.get("uco_nodes", {}).get("rows_error", 0) > 0:
            print(
                f"ERROR (strict): {report['uco_nodes']['rows_error']} uco_nodes rows had errors.",
                file=sys.stderr,
            )
            sys.exit(2)

    # ── 2. agency_registry ────────────────────────────────────
    ws_agency = find_sheet(wb, "agency_registry")
    if ws_agency is None:
        print("WARNING: Sheet 'agency_registry' not found; skipping.")
        report["agency_registry"] = {"skipped": True}
    else:
        print(f"[preprocess_workbook] Processing sheet: {ws_agency.title}")
        try:
            agency_data = process_agency_registry_sheet(ws_agency, report)
        except ValueError as exc:
            print(f"ERROR: agency_registry sheet validation failed: {exc}", file=sys.stderr)
            sys.exit(2)
        n = write_csv(
            os.path.join(output_dir, "agency_registry.csv"),
            agency_data,
            AGENCY_REGISTRY_CSV_COLS,
        )
        print(f"[preprocess_workbook] agency_registry.csv: {n} rows")

    # ── 3. naics_decoder ──────────────────────────────────────
    ws_naics = find_sheet(wb, "naics_decoder")
    if ws_naics is None:
        print("WARNING: Sheet 'naics_decoder' not found; skipping.")
        report["naics_decoder"] = {"skipped": True}
    else:
        print(f"[preprocess_workbook] Processing sheet: {ws_naics.title}")
        try:
            naics_data = process_naics_decoder_sheet(ws_naics, report)
        except ValueError as exc:
            print(f"ERROR: naics_decoder sheet validation failed: {exc}", file=sys.stderr)
            sys.exit(2)
        n = write_csv(
            os.path.join(output_dir, "naics_decoder.csv"),
            naics_data,
            NAICS_DECODER_CSV_COLS,
        )
        print(f"[preprocess_workbook] naics_decoder.csv: {n} rows")

    # ── 4. code_crosswalk ─────────────────────────────────────
    ws_crosswalk = find_sheet(wb, "code_crosswalk")
    if ws_crosswalk is None:
        print("WARNING: Sheet 'code_crosswalk' not found; skipping.")
        report["code_crosswalk"] = {"skipped": True}
    else:
        print(f"[preprocess_workbook] Processing sheet: {ws_crosswalk.title}")
        try:
            crosswalk_data = process_code_crosswalk_sheet(ws_crosswalk, report)
        except ValueError as exc:
            print(f"ERROR: code_crosswalk sheet validation failed: {exc}", file=sys.stderr)
            sys.exit(2)
        n = write_csv(
            os.path.join(output_dir, "code_crosswalk.csv"),
            crosswalk_data,
            CODE_CROSSWALK_CSV_COLS,
        )
        print(f"[preprocess_workbook] code_crosswalk.csv: {n} rows")

    # ── 5. obligation_metadata ────────────────────────────────
    ws_meta = find_sheet(wb, "obligation_metadata")
    if ws_meta is None:
        # Build metadata from uco_nodes jurisdiction normalization if no sheet
        print("INFO: Sheet '_obligation_metadata' not found; synthesizing from uco_nodes.")
        meta_data = _build_obligation_metadata_from_nodes(uco_nodes_data)
        report["obligation_metadata"] = {
            "synthesized": True,
            "rows_accepted": len(meta_data),
        }
    else:
        print(f"[preprocess_workbook] Processing sheet: {ws_meta.title}")
        try:
            meta_data = process_obligation_metadata_sheet(ws_meta, uco_nodes_data, report)
        except ValueError as exc:
            print(f"ERROR: _obligation_metadata sheet validation failed: {exc}", file=sys.stderr)
            sys.exit(2)

    # Merge in synthesized records for nodes not covered by explicit sheet rows
    meta_ids = {r["uco_node_id"] for r in meta_data}
    synthesized = _build_obligation_metadata_from_nodes(
        [n for n in uco_nodes_data if n["uco_node_id"] not in meta_ids]
    )
    if synthesized:
        print(f"[preprocess_workbook] Synthesizing {len(synthesized)} metadata records from jurisdiction normalization.")
        meta_data.extend(synthesized)

    n = write_csv(
        os.path.join(output_dir, "obligation_metadata.csv"),
        meta_data,
        OBLIGATION_METADATA_CSV_COLS,
    )
    print(f"[preprocess_workbook] obligation_metadata.csv: {n} rows")

    # ── 6. Transform report ───────────────────────────────────
    report_path = os.path.join(output_dir, "transform_report.json")
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)
    print(f"[preprocess_workbook] Transform report: {report_path}")

    # Summary
    total_errors = sum(
        v.get("rows_error", 0)
        for v in report.values()
        if isinstance(v, dict)
    )
    print(
        f"\n[preprocess_workbook] Done. "
        f"Nodes: {report.get('uco_nodes', {}).get('rows_accepted', 0)}, "
        f"Total errors: {total_errors}"
    )
    if total_errors > 0:
        print(
            f"WARNING: {total_errors} rows had errors. See {report_path} for details.",
            file=sys.stderr,
        )

    return report


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Preprocess UCO workbook into seed CSVs and a transform report.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        metavar="PATH",
        help="Path to source .xlsx workbook",
    )
    parser.add_argument(
        "--output-dir",
        default="db/seeds",
        metavar="DIR",
        help="Output directory for CSVs and transform_report.json (default: db/seeds)",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        default=False,
        help="Exit non-zero on any normalization error (default: warn and continue)",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"ERROR: --xlsx file not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    run_pipeline(args.xlsx, args.output_dir, strict=args.strict)


if __name__ == "__main__":
    main()
