#!/usr/bin/env python3
"""
tests/test_workbook_normalizer.py
IOS+ UCO Workbook Ingestion — Normalizer Unit Tests

Tests normalization functions and workbook processing behavior from
scripts/db/preprocess_workbook.py.

Run:
    pytest scripts/db/tests/test_workbook_normalizer.py -v
"""

import io
import json
import os
import sys
import csv
import tempfile
import unittest

# Make the scripts/db package importable from the repo root
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from preprocess_workbook import (
    normalize_ontology_level,
    normalize_ybr_gate,
    normalize_jurisdiction,
    _parse_jurisdiction,
    _build_obligation_metadata_from_nodes,
    process_obligation_metadata_sheet,
    VALID_ONTOLOGY_LEVELS,
    VALID_YBR_GATES,
    VALID_JURISDICTION_LEVELS,
    clean,
    header_index,
    read_sheet_with_headers,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_ws(headers: list, rows: list):
    """
    Build a minimal mock worksheet object that behaves like openpyxl's worksheet
    for iter_rows(values_only=True).
    """
    class MockWS:
        @property
        def title(self):
            return "test_sheet"

        def iter_rows(self, values_only=True):
            yield tuple(headers)
            for row in rows:
                yield tuple(row)

    return MockWS()


# ---------------------------------------------------------------------------
# 1. ontology_level normalization
# ---------------------------------------------------------------------------

class TestOntologyNormalization(unittest.TestCase):

    def test_canonical_values_pass_through(self):
        for val in VALID_ONTOLOGY_LEVELS:
            self.assertEqual(normalize_ontology_level(val), val)

    def test_sector_variants(self):
        for raw in ["sector", "SECTOR", "Sector", "l1: broad industry", "L1: Broad Industry", "broad industry"]:
            result = normalize_ontology_level(raw)
            self.assertEqual(result, "sector", f"Expected 'sector' for {raw!r}")

    def test_subsector_variants(self):
        for raw in [
            "subsector", "sub-sector", "sub sector", "industry subtype",
            "l2: regulations & rules", "L2: Regulations & Rules",
            "regulations & rules", "regulations and rules",
        ]:
            result = normalize_ontology_level(raw)
            self.assertEqual(result, "subsector", f"Expected 'subsector' for {raw!r}")

    def test_activity_variants(self):
        for raw in ["activity", "l3: specific activity", "specific activity", "functional"]:
            result = normalize_ontology_level(raw)
            self.assertEqual(result, "activity", f"Expected 'activity' for {raw!r}")

    def test_cross_cutting_variants(self):
        for raw in [
            "cross-cutting", "crosscutting", "cross cutting",
            "xsc", "xsc-cross-cutting",
        ]:
            result = normalize_ontology_level(raw)
            self.assertEqual(result, "cross-cutting", f"Expected 'cross-cutting' for {raw!r}")

    def test_invalid_raises(self):
        with self.assertRaises(ValueError):
            normalize_ontology_level("totally_unknown_value")

    def test_empty_raises(self):
        with self.assertRaises(ValueError):
            normalize_ontology_level(None)
        with self.assertRaises(ValueError):
            normalize_ontology_level("")

    def test_result_always_valid(self):
        """All successful normalizations must produce a valid DB value."""
        for raw in ["sector", "subsector", "activity", "cross-cutting", "functional", "xsc"]:
            result = normalize_ontology_level(raw)
            self.assertIn(result, VALID_ONTOLOGY_LEVELS)


# ---------------------------------------------------------------------------
# 2. ybr_gate normalization
# ---------------------------------------------------------------------------

class TestYBRGateNormalization(unittest.TestCase):

    def test_canonical_values_pass_through(self):
        for val in VALID_YBR_GATES:
            self.assertEqual(normalize_ybr_gate(val), val)

    def test_gate530_maps_to_L5(self):
        for raw in [
            "Gate 530: Compliance Check",
            "gate 530: compliance check",
            "Gate 530",
            "gate530",
            "Gate530: Compliance Check",
            "L5: Gate 530",
        ]:
            result = normalize_ybr_gate(raw)
            self.assertEqual(result, "L5", f"Expected 'L5' for {raw!r}")

    def test_l3_variants(self):
        for raw in ["l3", "L3", "l3: ontological mapping", "ontological mapping"]:
            result = normalize_ybr_gate(raw)
            self.assertEqual(result, "L3", f"Expected 'L3' for {raw!r}")

    def test_l4_variants(self):
        for raw in ["l4", "L4", "l4: evidence collection", "evidence collection"]:
            result = normalize_ybr_gate(raw)
            self.assertEqual(result, "L4", f"Expected 'L4' for {raw!r}")

    def test_l7_variants(self):
        for raw in ["l7", "L7", "l7: synthesis", "synthesis"]:
            result = normalize_ybr_gate(raw)
            self.assertEqual(result, "L7", f"Expected 'L7' for {raw!r}")

    def test_invalid_raises(self):
        with self.assertRaises(ValueError):
            normalize_ybr_gate("L99")

    def test_empty_raises(self):
        with self.assertRaises(ValueError):
            normalize_ybr_gate(None)
        with self.assertRaises(ValueError):
            normalize_ybr_gate("")

    def test_result_always_valid(self):
        for raw in ["l3", "l4", "l5", "l7", "Gate 530"]:
            result = normalize_ybr_gate(raw)
            self.assertIn(result, VALID_YBR_GATES)


# ---------------------------------------------------------------------------
# 3. jurisdiction normalization
# ---------------------------------------------------------------------------

class TestJurisdictionNormalization(unittest.TestCase):

    def test_canonical_values_pass_through(self):
        for val in VALID_JURISDICTION_LEVELS:
            level, state = normalize_jurisdiction(val)
            self.assertEqual(level, val)
            self.assertIsNone(state)

    def test_state_dash_tx(self):
        level, state = normalize_jurisdiction("State – TX")
        self.assertEqual(level, "State")
        self.assertEqual(state, "TX")

    def test_state_dash_ca(self):
        level, state = normalize_jurisdiction("State – CA")
        self.assertEqual(level, "State")
        self.assertEqual(state, "CA")

    def test_state_hyphen(self):
        level, state = normalize_jurisdiction("State - NY")
        self.assertEqual(level, "State")
        self.assertEqual(state, "NY")

    def test_state_emdash(self):
        level, state = normalize_jurisdiction("State — FL")
        self.assertEqual(level, "State")
        self.assertEqual(state, "FL")

    def test_state_paren(self):
        level, state = normalize_jurisdiction("State (TX)")
        self.assertEqual(level, "State")
        self.assertEqual(state, "TX")

    def test_federal_state_resolves_to_federal(self):
        level, state = normalize_jurisdiction("Federal / State")
        self.assertEqual(level, "Federal")
        self.assertIsNone(state)

    def test_federal_slash_state_no_spaces(self):
        level, state = normalize_jurisdiction("Federal/State")
        self.assertEqual(level, "Federal")
        self.assertIsNone(state)

    def test_state_local_resolves_to_state(self):
        level, state = normalize_jurisdiction("State / Local")
        self.assertEqual(level, "State")
        self.assertIsNone(state)

    def test_invalid_raises(self):
        with self.assertRaises(ValueError):
            normalize_jurisdiction("Galactic Empire")

    def test_empty_raises(self):
        with self.assertRaises(ValueError):
            normalize_jurisdiction(None)
        with self.assertRaises(ValueError):
            normalize_jurisdiction("")

    def test_result_always_valid_jurisdiction_level(self):
        cases = ["State – TX", "Federal", "State", "Local", "International", "Federal / State"]
        for raw in cases:
            level, _ = normalize_jurisdiction(raw)
            self.assertIn(level, VALID_JURISDICTION_LEVELS, f"Invalid DB value for {raw!r}")


# ---------------------------------------------------------------------------
# 4. Verification status / stale metadata handling
# ---------------------------------------------------------------------------

class TestVerificationMetadata(unittest.TestCase):

    def test_valid_verification_statuses(self):
        valid = {"verified", "stale", "corrected", "pending"}
        ws = _make_ws(
            ["uco_node_id", "verification_status"],
            [
                ["UCO-TEST-001", "verified"],
                ["UCO-TEST-002", "stale"],
                ["UCO-TEST-003", "corrected"],
                ["UCO-TEST-004", "pending"],
            ],
        )
        report = {}
        results = process_obligation_metadata_sheet(ws, [], report)
        statuses = {r["uco_node_id"]: r["verification_status"] for r in results}
        self.assertEqual(statuses["UCO-TEST-001"], "verified")
        self.assertEqual(statuses["UCO-TEST-002"], "stale")
        self.assertEqual(statuses["UCO-TEST-003"], "corrected")
        self.assertEqual(statuses["UCO-TEST-004"], "pending")

    def test_unknown_status_defaults_to_pending(self):
        ws = _make_ws(
            ["uco_node_id", "verification_status"],
            [["UCO-UNKNOWN-001", "outdated"]],
        )
        report = {}
        results = process_obligation_metadata_sheet(ws, [], report)
        self.assertEqual(results[0]["verification_status"], "pending")
        # Should record a warning
        self.assertGreater(len(report["obligation_metadata"]["errors"]), 0)

    def test_stale_node_preserves_provenance(self):
        """A stale record keeps its jurisdiction_detail for audit purposes."""
        ws = _make_ws(
            ["uco_node_id", "verification_status", "jurisdiction_detail", "source_note"],
            [["UCO-STALE-001", "stale", "State – TX", "Needs re-verification by 2026-Q3"]],
        )
        report = {}
        results = process_obligation_metadata_sheet(ws, [], report)
        self.assertEqual(results[0]["verification_status"], "stale")
        self.assertEqual(results[0]["jurisdiction_detail"], "State – TX")
        self.assertEqual(results[0]["source_note"], "Needs re-verification by 2026-Q3")

    def test_corrected_record_has_explicit_note(self):
        ws = _make_ws(
            ["uco_node_id", "verification_status", "source_note"],
            [["UCO-CORRECTED-001", "corrected", "Was previously marked Federal; corrected to State"]],
        )
        report = {}
        results = process_obligation_metadata_sheet(ws, [], report)
        self.assertEqual(results[0]["verification_status"], "corrected")
        self.assertIn("corrected", results[0]["source_note"].lower())

    def test_as_of_date_parsed_correctly(self):
        ws = _make_ws(
            ["uco_node_id", "as_of_date"],
            [
                ["UCO-DATE-001", "2026-01-15"],
                ["UCO-DATE-002", "01/15/2026"],
            ],
        )
        report = {}
        results = process_obligation_metadata_sheet(ws, [], report)
        dates = {r["uco_node_id"]: r["as_of_date"] for r in results}
        self.assertEqual(dates["UCO-DATE-001"], "2026-01-15")
        self.assertEqual(dates["UCO-DATE-002"], "2026-01-15")


# ---------------------------------------------------------------------------
# 5. Multi-sheet workbook processing behavior
# ---------------------------------------------------------------------------

class TestMultiSheetProcessing(unittest.TestCase):

    def _make_workbook(self, sheet_specs: dict) -> "openpyxl.Workbook":
        """
        Build an openpyxl Workbook with specified sheets.
        sheet_specs = { sheet_name: (headers, rows) }
        """
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        wb = openpyxl.Workbook()
        first = True
        for name, (headers, rows) in sheet_specs.items():
            if first:
                ws = wb.active
                ws.title = name
                first = False
            else:
                ws = wb.create_sheet(name)
            ws.append(headers)
            for row in rows:
                ws.append(row)
        return wb

    def test_uco_nodes_sheet_processed_correctly(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        from preprocess_workbook import process_uco_nodes_sheet

        ws_data = _make_ws(
            [
                "broad_industry", "industry_subtype", "specific_activity",
                "jurisdiction_level", "governing_agency", "regulation_name",
                "cfr_usc_citation", "report_form_name", "form_code", "filing_frequency",
                "key_due_dates", "business_segment", "penalties_consequences",
                "cip", "sic", "naics", "soc", "isic", "hs_hts", "notes",
                "uco_node_id", "ontology_level", "compliance_chain_ref", "operating_segment",
                "responsible_role", "enforcement_type", "risk_weight", "ybr_gate",
                "policy_action", "last_updated",
            ],
            [
                [
                    "Energy", "Oil & Gas", "Exploration", "Federal", "EPA",
                    "Clean Air Act", "42 U.S.C. § 7401", "Form CAA-1", "CAA-1",
                    "Annual", "March 31", "Operations", "$50,000/day", None, "2911",
                    "211111", None, None, None, None,
                    "UCO-ENERGY-001", "sector", None, "Operations",
                    "Environmental Officer", "Civil Monetary Penalty", "8",
                    "L5", "BLOCK", "2026-01-01",
                ]
            ],
        )
        report = {}
        results = process_uco_nodes_sheet(ws_data, report)
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["uco_node_id"], "UCO-ENERGY-001")
        self.assertEqual(results[0]["ontology_level"], "sector")
        self.assertEqual(results[0]["ybr_gate"], "L5")
        self.assertEqual(results[0]["jurisdiction_level"], "Federal")

    def test_jurisdiction_detail_preserved_in_metadata(self):
        """Nodes with non-canonical jurisdiction should yield metadata records."""
        uco_nodes = [
            {
                "uco_node_id": "UCO-STATE-001",
                "jurisdiction_level": "State",
                "_jurisdiction_detail": "State – TX",
                "_state_code": "TX",
            },
            {
                "uco_node_id": "UCO-FED-001",
                "jurisdiction_level": "Federal",
                "_jurisdiction_detail": "Federal",
                "_state_code": None,
            },
        ]
        meta = _build_obligation_metadata_from_nodes(uco_nodes)
        # Only UCO-STATE-001 has a different detail (TX) → should produce metadata
        ids = [m["uco_node_id"] for m in meta]
        self.assertIn("UCO-STATE-001", ids)
        tx_record = next(m for m in meta if m["uco_node_id"] == "UCO-STATE-001")
        self.assertEqual(tx_record["jurisdiction_detail"], "State – TX")
        self.assertEqual(tx_record["state"], "TX")
        # Federal canonical has no discrepancy → may or may not appear; if it does,
        # it should not have a TX state code
        fed_records = [m for m in meta if m["uco_node_id"] == "UCO-FED-001"]
        if fed_records:
            self.assertIsNone(fed_records[0]["state"])

    def test_obligation_metadata_merges_jurisdiction_from_nodes(self):
        """Metadata sheet records should be enriched with state from uco_nodes."""
        # uco_nodes data carries _jurisdiction_detail and _state_code
        uco_nodes = [
            {
                "uco_node_id": "UCO-TX-001",
                "jurisdiction_level": "State",
                "_jurisdiction_detail": "State – TX",
                "_state_code": "TX",
            }
        ]
        # Sheet does NOT include jurisdiction_detail or state columns
        ws = _make_ws(
            ["uco_node_id", "verification_status"],
            [["UCO-TX-001", "verified"]],
        )
        report = {}
        results = process_obligation_metadata_sheet(ws, uco_nodes, report)
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["jurisdiction_detail"], "State – TX")
        self.assertEqual(results[0]["state"], "TX")

    def test_empty_sheet_returns_empty(self):
        ws = _make_ws(["uco_node_id"], [])
        report = {}
        results = process_obligation_metadata_sheet(ws, [], report)
        self.assertEqual(results, [])

    def test_missing_required_column_raises(self):
        from preprocess_workbook import _validate_required_cols
        headers = ["some_other_col", "another_col"]
        with self.assertRaises(ValueError) as ctx:
            _validate_required_cols(headers, ["uco_node_id"], "uco_nodes")
        self.assertIn("uco_node_id", str(ctx.exception))

    def test_row_deduplication_in_uco_nodes(self):
        """Duplicate uco_node_id rows should be skipped after first occurrence."""
        ws = _make_ws(
            [
                "broad_industry", "industry_subtype", "specific_activity",
                "jurisdiction_level", "governing_agency", "regulation_name",
                "cfr_usc_citation", "report_form_name", "form_code", "filing_frequency",
                "key_due_dates", "business_segment", "penalties_consequences",
                "cip", "sic", "naics", "soc", "isic", "hs_hts", "notes",
                "uco_node_id", "ontology_level", "compliance_chain_ref", "operating_segment",
                "responsible_role", "enforcement_type", "risk_weight", "ybr_gate",
                "policy_action", "last_updated",
            ],
            [
                # Row 1
                [
                    "Energy", "Oil & Gas", "Exploration", "Federal", "EPA",
                    "Clean Air Act", None, None, None, None, None, None, None,
                    None, None, "211111", None, None, None, None,
                    "UCO-ENERGY-001", "sector", None, None,
                    None, "Civil Monetary Penalty", "8", "L5", "BLOCK", None,
                ],
                # Row 2 — duplicate ID
                [
                    "Energy", "Oil & Gas", "Exploration v2", "Federal", "EPA",
                    "Clean Air Act v2", None, None, None, None, None, None, None,
                    None, None, "211111", None, None, None, None,
                    "UCO-ENERGY-001", "sector", None, None,
                    None, "Civil Monetary Penalty", "8", "L5", "BLOCK", None,
                ],
            ],
        )
        from preprocess_workbook import process_uco_nodes_sheet
        report = {}
        results = process_uco_nodes_sheet(ws, report)
        ids = [r["uco_node_id"] for r in results]
        self.assertEqual(ids.count("UCO-ENERGY-001"), 1)


# ---------------------------------------------------------------------------
# 6. Clean helper tests
# ---------------------------------------------------------------------------

class TestCleanHelper(unittest.TestCase):

    def test_none_returns_none(self):
        self.assertIsNone(clean(None))

    def test_empty_string_returns_none(self):
        self.assertIsNone(clean(""))
        self.assertIsNone(clean("   "))

    def test_none_string_returns_none(self):
        self.assertIsNone(clean("none"))
        self.assertIsNone(clean("None"))
        self.assertIsNone(clean("NONE"))

    def test_na_returns_none(self):
        self.assertIsNone(clean("n/a"))
        self.assertIsNone(clean("N/A"))
        self.assertIsNone(clean("na"))
        self.assertIsNone(clean("#N/A"))

    def test_valid_value_strips_whitespace(self):
        self.assertEqual(clean("  hello  "), "hello")
        self.assertEqual(clean("EPA"), "EPA")


# ---------------------------------------------------------------------------
# 7. header_index helper tests
# ---------------------------------------------------------------------------

class TestHeaderIndex(unittest.TestCase):

    def test_finds_exact_match(self):
        idx = header_index(["uco_node_id", "naics", "jurisdiction"], "naics")
        self.assertEqual(idx, 1)

    def test_case_insensitive(self):
        idx = header_index(["UCO_Node_ID", "NAICS", "Jurisdiction"], "uco_node_id")
        self.assertEqual(idx, 0)

    def test_missing_required_raises(self):
        with self.assertRaises(ValueError):
            header_index(["a", "b"], "c", required=True)

    def test_missing_optional_returns_minus_one(self):
        idx = header_index(["a", "b"], "c", required=False)
        self.assertEqual(idx, -1)


if __name__ == "__main__":
    unittest.main(verbosity=2)
